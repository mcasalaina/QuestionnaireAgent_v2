#!/usr/bin/env python3
"""
Questionnaire Multiagent - UI Application

A windowed application that orchestrates three Azure AI Foundry agents to answer questions
with fact-checking and link validation. Supports both individual questions and Excel import/export.
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import os
import tempfile
import asyncio
import logging
import argparse
import sys
from typing import Tuple, Optional, List, Dict, Any
from pathlib import Path
from azure.monitor.opentelemetry import configure_azure_monitor
from opentelemetry import trace
from opentelemetry.trace import Tracer

import pandas as pd
from azure.ai.projects import AIProjectClient
from azure.identity import DefaultAzureCredential
from azure.ai.agents.models import BingGroundingTool
from dotenv import load_dotenv

# Load environment variables
load_dotenv(override=True)

class QuestionnaireAgentUI:
    """Main UI application for the Questionnaire Agent."""
    
    def __init__(self, headless_mode=False, max_retries=10):
        # Setup logging first - only for our app, not Azure SDK noise
        logging.basicConfig(level=logging.WARNING)  # Set root to WARNING to silence Azure SDK
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)  # But keep our app logger at INFO
        
        # Silence specific noisy Azure loggers
        logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)
        logging.getLogger("azure.monitor.opentelemetry.exporter").setLevel(logging.WARNING)
        logging.getLogger("azure.monitor.opentelemetry").setLevel(logging.INFO)
        
        # Store headless mode flag
        self.headless_mode = headless_mode
        
        # Store maximum retries configuration
        self.max_retries = max_retries
        
        if not headless_mode:
            # Enable high DPI awareness on Windows
            try:
                from ctypes import windll
                windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass  # Ignore errors on non-Windows platforms or if not available
            
            self.root = tk.Tk()
            self.root.title("Questionnaire Multiagent")
            self.root.geometry("1200x800")
            self.root.state('zoomed')  # Open maximized on Windows
        else:
            self.root = None
        
        # Initialize Azure AI Project Client
        self.project_client = None
        self.init_azure_client()
        
        # Agent IDs will be set when agents are created
        self.question_answerer_id = None
        self.answer_checker_id = None
        self.link_checker_id = None
        
        # CLI output buffer for reasoning
        self.cli_output = []
        
        # OpenTelemetry tracer for Azure AI Foundry tracing
        self.tracer: Optional[Tracer] = None
        
        # Initialize tracing
        self.initialize_tracing()
        
        # Setup UI only if not in headless mode
        if not headless_mode:
            self.setup_ui()
        
    def init_azure_client(self):
        """Initialize Azure AI Project Client with credentials from .env file."""
        try:
            endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
            if not endpoint:
                raise ValueError("AZURE_OPENAI_ENDPOINT not found in environment variables")
            
            self.logger.info(f"Connecting to Azure AI Foundry endpoint: {endpoint}")
            
            credential = DefaultAzureCredential()
            self.project_client = AIProjectClient(
                endpoint=endpoint,
                credential=credential
            )
            self.logger.info("Azure AI Project Client initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Failed to initialize Azure client: {e}")
            if not self.headless_mode:
                messagebox.showerror("Azure Connection Error", 
                                   f"Failed to connect to Azure AI Foundry:\n{e}\n\nPlease check your .env file and Azure credentials.")
            else:
                print(f"Error: Failed to connect to Azure AI Foundry: {e}")
                print("Please check your .env file and Azure credentials.")
                sys.exit(1)
    
    def initialize_tracing(self):
        """Initialize Azure AI Foundry tracing with Application Insights"""
        try:
            # Configure content recording based on environment variable or default to true for debugging
            content_recording = os.environ.get("AZURE_TRACING_GEN_AI_CONTENT_RECORDING_ENABLED", "true").lower()
            os.environ["AZURE_TRACING_GEN_AI_CONTENT_RECORDING_ENABLED"] = content_recording
            
            # Get Application Insights connection string from environment variable
            connection_string = os.environ.get("APPLICATIONINSIGHTS_CONNECTION_STRING")
            
            if not connection_string:
                self.logger.warning("⚠️ APPLICATIONINSIGHTS_CONNECTION_STRING environment variable not set.")
                self.logger.warning("   Set this in your .env file - get it from Azure Portal > Application Insights > Overview")
                return False
            
            # Configure Azure Monitor tracing
            configure_azure_monitor(connection_string=connection_string)
            
            # Create a tracer for custom spans
            self.tracer = trace.get_tracer(__name__)
            
            content_status = "enabled" if content_recording == "true" else "disabled"
            self.logger.info(f"✅ Azure AI Foundry tracing initialized successfully (content recording: {content_status}).")
            return True
            
        except Exception as e:
            self.logger.warning(f"⚠️ Failed to initialize tracing: {str(e)}")
            return False
    
    def setup_ui(self):
        """Setup the main UI layout."""
        # Create main paned window
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)
        
        # Right panel
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=2)
        
        self.setup_left_panel(left_frame)
        self.setup_right_panel(right_frame)
        
    def setup_left_panel(self, parent):
        """Setup the left panel with input controls."""
        # Context section
        context_label = ttk.Label(parent, text="Context")
        context_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.context_entry = tk.Entry(parent, width=40)
        self.context_entry.insert(0, "Microsoft Azure AI")  # Default value
        self.context_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Character Limit section
        limit_label = ttk.Label(parent, text="Character Limit")
        limit_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.limit_entry = tk.Entry(parent, width=40)
        self.limit_entry.insert(0, "2000")  # Default value
        self.limit_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Maximum Retries section
        retries_label = ttk.Label(parent, text="Maximum Retries")
        retries_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.retries_entry = tk.Entry(parent, width=40)
        self.retries_entry.insert(0, str(self.max_retries))  # Default value from constructor
        self.retries_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Question section
        question_label = ttk.Label(parent, text="Question")
        question_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.question_text = scrolledtext.ScrolledText(parent, height=8, width=40, 
                                                     font=('Segoe UI', 12))
        self.question_text.insert(tk.END, "Does your service offer video generative AI?")  # Default value
        self.question_text.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.ask_button = tk.Button(button_frame, text="Ask!", bg="lightgray", fg="black", 
                                  height=2, command=self.on_ask_clicked)
        self.ask_button.pack(fill=tk.X, pady=(0, 10))
        
        self.import_button = tk.Button(button_frame, text="📊 Import From Excel", 
                                     command=self.on_import_excel_clicked)
        self.import_button.pack(fill=tk.X)
        
    def setup_right_panel(self, parent):
        """Setup the right panel with output sections."""
        # Create notebook for tabbed interface
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Answer tab
        answer_frame = ttk.Frame(self.notebook)
        self.notebook.add(answer_frame, text="Answer")
        
        answer_label = ttk.Label(answer_frame, text="Answer")
        answer_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.answer_text = scrolledtext.ScrolledText(answer_frame, wrap=tk.WORD, 
                                                   font=('Segoe UI', 12))
        self.answer_text.insert(tk.END, "Response will appear here after clicking Ask!")
        self.answer_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
        # Documentation tab
        docs_frame = ttk.Frame(self.notebook)
        self.notebook.add(docs_frame, text="Documentation")
        
        docs_label = ttk.Label(docs_frame, text="Documentation")
        docs_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.docs_text = scrolledtext.ScrolledText(docs_frame, wrap=tk.WORD, 
                                                  font=('Segoe UI', 12))
        self.docs_text.insert(tk.END, "Documentation will appear here...")
        self.docs_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
        # Reasoning tab
        reasoning_frame = ttk.Frame(self.notebook)
        self.notebook.add(reasoning_frame, text="Reasoning")
        
        reasoning_label = ttk.Label(reasoning_frame, text="Reasoning")
        reasoning_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.reasoning_text = scrolledtext.ScrolledText(reasoning_frame, wrap=tk.WORD, 
                                                      font=('Segoe UI', 12))
        self.reasoning_text.insert(tk.END, "Reasoning will appear here...")
        self.reasoning_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
    def log_reasoning(self, message: str):
        """Add a message to the reasoning text area or CLI output."""
        if self.headless_mode:
            self.cli_output.append(message)
        else:
            self.reasoning_text.insert(tk.END, f"{message}\n")
            self.reasoning_text.see(tk.END)
            self.root.update_idletasks()
        
    def on_ask_clicked(self):
        """Handle the Ask button click."""
        # Disable button during processing
        self.ask_button.config(state=tk.DISABLED)
        
        # Switch to Reasoning tab immediately
        self.notebook.select(2)  # Index 2 is the Reasoning tab (Answer=0, Documentation=1, Reasoning=2)
        
        # Clear previous results
        self.answer_text.delete(1.0, tk.END)
        self.docs_text.delete(1.0, tk.END)
        self.reasoning_text.delete(1.0, tk.END)
        
        # Get input values
        question = self.question_text.get(1.0, tk.END).strip()
        context = self.context_entry.get().strip()
        char_limit = int(self.limit_entry.get()) if self.limit_entry.get().isdigit() else 2000
        max_retries = int(self.retries_entry.get()) if self.retries_entry.get().isdigit() else self.max_retries
        
        if not question:
            messagebox.showwarning("Input Required", "Please enter a question.")
            self.ask_button.config(state=tk.NORMAL)
            return
            
        # Run processing in separate thread
        thread = threading.Thread(target=self.process_single_question, 
                                args=(question, context, char_limit, max_retries))
        thread.daemon = True
        thread.start()
        
    def process_single_question(self, question: str, context: str, char_limit: int, max_retries: int):
        """Process a single question using the three-agent workflow."""
        try:
            # Use custom span for the entire workflow
            if self.tracer:
                with self.tracer.start_as_current_span("questionnaire_multi_agent_workflow") as span:
                    span.set_attribute("workflow.name", "Questionnaire Multi-Agent")
                    span.set_attribute("workflow.context", context)
                    span.set_attribute("workflow.char_limit", char_limit)
                    span.set_attribute("workflow.max_retries", max_retries)
                    span.set_attribute("question.text", question[:100] + "..." if len(question) > 100 else question)
                    success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
                    return success, answer, links
            else:
                success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
                return success, answer, links
                
        except Exception as e:
            self.logger.error(f"Error processing question: {e}")
            error_msg = str(e)  # Capture the error message as a string
            if not self.headless_mode:
                self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {error_msg}"))
            return False, f"Error: {error_msg}", []
        finally:
            # Re-enable button
            if not self.headless_mode:
                self.root.after(0, lambda: self.ask_button.config(state=tk.NORMAL))
    
    def _execute_workflow(self, question: str, context: str, char_limit: int, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Internal method to execute the multi-agent workflow."""
        self.log_reasoning("Starting question processing...")
        
        # Create agents if not already created
        if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
            self.create_agents()
        
        attempt = 1
        max_attempts = max_retries
        
        # Track valid links across retries for this question
        accumulated_valid_links = []
        
        while attempt <= max_attempts:
            self.log_reasoning(f"Attempt {attempt}/{max_attempts}")
            
            # Step 1: Generate answer
            self.log_reasoning("Question Answerer: Generating answer...")
            candidate_answer, doc_urls = self.generate_answer(question, context, char_limit)
            
            if not candidate_answer:
                self.log_reasoning("Question Answerer failed to generate an answer")
                break
            
            # Log the raw answer before processing
            self.log_reasoning("=== RAW ANSWER FROM QUESTION ANSWERER ===")
            self.log_reasoning(candidate_answer)
            self.log_reasoning("=== END RAW ANSWER ===")
            
            # Remove links and citations, save links for documentation
            clean_answer, text_links = self.extract_links_and_clean(candidate_answer)
            
            self.log_reasoning("=== CLEANED ANSWER AFTER URL EXTRACTION ===")
            self.log_reasoning(clean_answer)
            self.log_reasoning("=== END CLEANED ANSWER ===")
            
            self.log_reasoning(f"Extracted {len(text_links)} URLs from answer text")
            for link in text_links:
                self.log_reasoning(f"  Text URL: {link}")
            
            # Combine documentation URLs from run steps with any URLs found in text
            all_links = list(set(doc_urls + text_links))  # Remove duplicates
            self.log_reasoning(f"Total combined URLs: {len(all_links)}")
            for link in all_links:
                self.log_reasoning(f"  Combined URL: {link}")
            
            # Check character limit
            if len(clean_answer) > char_limit:
                self.log_reasoning(f"Answer exceeds character limit ({len(clean_answer)} > {char_limit}). Retrying...")
                attempt += 1
                continue
            
            # Step 2: Validate answer
            self.log_reasoning("Answer Checker: Validating answer...")
            answer_valid, answer_feedback = self.validate_answer(question, clean_answer)
            
            if not answer_valid:
                self.log_reasoning(f"Answer Checker rejected: {answer_feedback}")
                attempt += 1
                continue
            
            # Step 3: Validate links
            self.log_reasoning("Link Checker: Verifying URLs...")
            links_valid, valid_links, link_feedback = self.validate_links(all_links)
            
            # Accumulate any valid links found in this attempt
            if valid_links:
                for link in valid_links:
                    if link not in accumulated_valid_links:
                        accumulated_valid_links.append(link)
                        self.log_reasoning(f"Added valid link to accumulated collection: {link}")
            
            # Check if we have a valid answer and at least some valid links (current or accumulated)
            if not links_valid and not accumulated_valid_links:
                # No valid links in current attempt and no accumulated links from previous attempts
                self.log_reasoning(f"Link Checker rejected: {link_feedback}")
                attempt += 1
                continue
            elif not links_valid and accumulated_valid_links:
                # Current attempt has no valid links, but we have accumulated valid links from previous attempts
                self.log_reasoning(f"Link Checker found no valid links in current attempt, but reusing {len(accumulated_valid_links)} valid links from previous attempts")
                final_valid_links = accumulated_valid_links.copy()
            else:
                # Current attempt has valid links
                final_valid_links = accumulated_valid_links.copy()  # Use all accumulated links
            
            # All checks passed
            self.log_reasoning("All agents approved the answer!")
            self.log_reasoning(f"Final answer will use {len(final_valid_links)} documentation links")
            
            # Update UI with results (only in non-headless mode)
            if not self.headless_mode:
                self.root.after(0, lambda: self.update_results(clean_answer, final_valid_links))
            
            # Return success with results
            return True, clean_answer, final_valid_links
            
        # Max attempts reached
        self.log_reasoning(f"Failed to generate acceptable answer after {max_attempts} attempts")
        if not self.headless_mode:
            self.root.after(0, lambda: messagebox.showerror("Processing Failed", 
                f"Could not generate an acceptable answer after {max_attempts} attempts."))
        
        return False, f"Failed to generate acceptable answer after {max_attempts} attempts", []
                
    def update_results(self, answer: str, links: List[str]):
        """Update the UI with the final answer and documentation."""
        self.answer_text.delete(1.0, tk.END)
        self.answer_text.insert(tk.END, answer)
        
        self.docs_text.delete(1.0, tk.END)
        if links:
            self.docs_text.insert(tk.END, "Documentation links:\n\n")
            for link in links:
                self.docs_text.insert(tk.END, f"• {link}\n")
        else:
            self.docs_text.insert(tk.END, "No documentation links found.")
            
    def create_agents(self):
        """Create the three Azure AI Foundry agents."""
        try:
            self.log_reasoning("Creating Azure AI Foundry agents...")
            
            # Get model deployment name from environment
            model_deployment = os.getenv("AZURE_OPENAI_MODEL_DEPLOYMENT")
            bing_resource_name = os.getenv("BING_CONNECTION_ID")
            
            self.log_reasoning(f"Using model deployment: {model_deployment}")
            self.log_reasoning(f"Using Bing connection name: {bing_resource_name}")
            
            if not bing_resource_name:
                raise ValueError("BING_CONNECTION_ID not found in environment variables")
            
            # Get the actual connection ID from the resource name
            self.log_reasoning(f"Getting connection ID for resource: {bing_resource_name}")
            try:
                connection = self.project_client.connections.get(name=bing_resource_name)
                conn_id = connection.id
                self.log_reasoning(f"Retrieved connection ID: {conn_id}")
            except Exception as conn_error:
                self.log_reasoning(f"ERROR: Failed to get Bing connection '{bing_resource_name}': {conn_error}")
                
                # List all available connections to help debug
                try:
                    self.log_reasoning("Listing all available connections:")
                    connections = self.project_client.connections.list()
                    for conn in connections:
                        self.log_reasoning(f"  - Name: '{conn.name}', Type: {getattr(conn, 'connection_type', 'unknown')}, ID: {conn.id}")
                    
                    if not connections:
                        self.log_reasoning("  No connections found in this project.")
                except Exception as list_error:
                    self.log_reasoning(f"  Could not list connections: {list_error}")
                
                raise ValueError(f"Bing connection '{bing_resource_name}' not found. Check BING_CONNECTION_ID in your .env file.") from conn_error
            
            # Create Bing grounding tool
            bing_tool = BingGroundingTool(connection_id=conn_id)
            
            # Create Question Answerer agent
            self.log_reasoning(f"Creating Question Answerer agent with model: {model_deployment}")
            try:
                question_answerer = self.project_client.agents.create_agent(
                    model=model_deployment,
                    name="Question Answerer",
                    instructions="You are a question answering agent. You MUST search the web extensively for evidence and synthesize accurate answers. Your answer must be based on current web search results. IMPORTANT: You must include the actual source URLs directly in your answer text. Write the full URLs (like https://docs.microsoft.com/example) in your response text where you reference information. Do not use citation markers like [1], (source), or 【†source】 - instead include the actual URLs, which you should always put at the end of your response, separated by newlines with no other text or formatting. Write in plain text without formatting. Your answer must end with a period and contain only complete sentences. Do not include any closing phrases like 'Learn more:', 'References:', 'For more information, see:', 'For more details, see:', 'Learn more at:', 'More information:', 'Additional resources:', or any similar calls-to-action at the end. There should only be prose, followed by a list of URLs for reference separated by newlines. Those URLs should be the ones provided by Bing. Always use the Bing grounding tool to search for current information.",
                    tools=bing_tool.definitions
                )
                self.question_answerer_id = question_answerer.id
                self.log_reasoning(f"Created Question Answerer agent: {self.question_answerer_id}")
            except Exception as e:
                self.log_reasoning(f"ERROR: Failed to create Question Answerer agent with model '{model_deployment}': {e}")
                raise ValueError(f"Model deployment '{model_deployment}' not found. Check AZURE_OPENAI_MODEL_DEPLOYMENT in your .env file.") from e
            
            # Create Answer Checker agent
            answer_checker = self.project_client.agents.create_agent(
                model=model_deployment,
                name="Answer Checker",
                instructions="You are an answer validation agent. Review candidate answers for factual correctness, completeness, and consistency. Use web search to verify claims. Respond with 'VALID' if the answer is acceptable or 'INVALID: [reason]' if not.",
                tools=bing_tool.definitions
            )
            self.answer_checker_id = answer_checker.id
            
            # Create Link Checker agent
            link_checker = self.project_client.agents.create_agent(
                model=model_deployment,
                name="Link Checker",
                instructions="You are a link validation agent. Verify that URLs are reachable and relevant to the given question. Report any issues with links.",
                tools=[]  # Will use requests/playwright for link checking
            )
            self.link_checker_id = link_checker.id
            
            self.log_reasoning("All agents created successfully!")
            
        except Exception as e:
            self.logger.error(f"Failed to create agents: {e}")
            raise
            
    def generate_answer(self, question: str, context: str, char_limit: int) -> Tuple[Optional[str], List[str]]:
        """Generate an answer using the Question Answerer agent."""
        try:
            # Use custom span for Question Answerer agent
            if self.tracer:
                with self.tracer.start_as_current_span("question_answerer_agent") as span:
                    span.set_attribute("agent.name", "Question Answerer")
                    span.set_attribute("agent.operation", "generate_answer")
                    span.set_attribute("question.context", context)
                    span.set_attribute("question.char_limit", char_limit)
                    return self._execute_question_answerer(question, context, char_limit)
            else:
                return self._execute_question_answerer(question, context, char_limit)
                
        except Exception as e:
            error_msg = f"Error generating answer: {e}"
            self.logger.error(error_msg)
            self.log_reasoning(error_msg)
            return None, []
    
    def _execute_question_answerer(self, question: str, context: str, char_limit: int) -> Tuple[Optional[str], List[str]]:
        """Internal method to execute Question Answerer agent operations."""
        # Create thread
        thread = self.project_client.agents.threads.create()
        self.log_reasoning(f"Created thread: {thread.id}")
        
        # Create message
        message = self.project_client.agents.messages.create(
            thread_id=thread.id,
            role="user",
            content=f"Context: {context}\n\nQuestion: {question}\n\nPlease provide a comprehensive answer with supporting evidence and citations. Keep it under {char_limit} characters."
        )
        self.log_reasoning(f"Created message: {message.id}")
        
        # Create and process run
        run = self.project_client.agents.runs.create_and_process(
            thread_id=thread.id,
            agent_id=self.question_answerer_id
        )
        
        self.log_reasoning(f"Run finished with status: {run.status}")
        
        # Check if the run failed
        if run.status == "failed":
            error_msg = f"Run failed: {run.last_error.message if run.last_error else 'Unknown error'}"
            self.log_reasoning(error_msg)
            self.logger.error(error_msg)
            return None, []
        
        # Get messages
        messages = self.project_client.agents.messages.list(thread_id=thread.id)
        
        # Find the assistant's response
        for msg in messages:
            if msg.role == "assistant" and msg.content:
                response = msg.content[0].text.value
                self.log_reasoning(f"Got response: {response[:100]}...")
                
                # Extract actual source URLs from message annotations
                doc_urls = []
                if hasattr(msg.content[0], 'annotations') and msg.content[0].annotations:
                    self.log_reasoning(f"Found {len(msg.content[0].annotations)} annotations")
                    for annotation in msg.content[0].annotations:
                        if hasattr(annotation, 'uri_citation') and annotation.uri_citation:
                            url = annotation.uri_citation.uri
                            # Only include actual website URLs, not Bing API URLs
                            if url and not url.startswith('https://api.bing.microsoft.com'):
                                doc_urls.append(url)
                                self.log_reasoning(f"Found source URL: {url}")
                
                if not doc_urls:
                    self.log_reasoning("No annotations found, trying run steps...")
                    doc_urls = self.extract_documentation_urls(thread.id, run.id)
                
                return response, doc_urls
                
        self.log_reasoning("No assistant response found in messages")
        return None, []
            
    def extract_documentation_urls(self, thread_id: str, run_id: str) -> List[str]:
        """Extract documentation URLs from run steps."""
        try:
            run_steps = self.project_client.agents.run_steps.list(thread_id=thread_id, run_id=run_id)
            documentation_urls = []
            
            for step in run_steps:
                self.log_reasoning(f"Checking run step: {step.id}")
                
                # Check if there are tool calls in the step details
                if hasattr(step, 'step_details') and step.step_details:
                    step_details = step.step_details
                    if hasattr(step_details, 'tool_calls') and step_details.tool_calls:
                        self.log_reasoning(f"Found {len(step_details.tool_calls)} tool calls")
                        
                        for call in step_details.tool_calls:
                            # Look for any tool call output that might contain URLs
                            if hasattr(call, 'bing_grounding') and call.bing_grounding:
                                self.log_reasoning(f"Examining bing_grounding data: {type(call.bing_grounding)}")
                                
                                # Handle different data structures
                                if isinstance(call.bing_grounding, dict):
                                    for key, value in call.bing_grounding.items():
                                        self.log_reasoning(f"  Key: {key}, Value type: {type(value)}")
                                        if isinstance(value, str) and value.startswith('http') and not value.startswith('https://api.bing.microsoft.com'):
                                            documentation_urls.append(value)
                                            self.log_reasoning(f"Found documentation URL from dict: {value}")
                                elif isinstance(call.bing_grounding, list):
                                    for item in call.bing_grounding:
                                        if isinstance(item, dict):
                                            for key, value in item.items():
                                                if isinstance(value, str) and value.startswith('http') and not value.startswith('https://api.bing.microsoft.com'):
                                                    documentation_urls.append(value)
                                                    self.log_reasoning(f"Found documentation URL from list: {value}")
                                        elif isinstance(item, str) and item.startswith('http') and not item.startswith('https://api.bing.microsoft.com'):
                                            documentation_urls.append(item)
                                            self.log_reasoning(f"Found documentation URL from list item: {item}")
            
            if documentation_urls:
                self.log_reasoning(f"Extracted {len(documentation_urls)} documentation URLs")
            else:
                self.log_reasoning("No documentation URLs found in run steps")
                
            return list(set(documentation_urls))  # Remove duplicates
            
        except Exception as e:
            self.log_reasoning(f"Error extracting documentation URLs: {e}")
            return []
            
    def validate_answer(self, question: str, answer: str) -> Tuple[bool, str]:
        """Validate an answer using the Answer Checker agent."""
        try:
            # Use custom span for Answer Checker agent
            if self.tracer:
                with self.tracer.start_as_current_span("answer_checker_agent") as span:
                    span.set_attribute("agent.name", "Answer Checker")
                    span.set_attribute("agent.operation", "validate_answer")
                    span.set_attribute("answer.length", len(answer))
                    return self._execute_answer_checker(question, answer)
            else:
                return self._execute_answer_checker(question, answer)
                
        except Exception as e:
            self.logger.error(f"Error validating answer: {e}")
            return False, f"Error: {e}"
    
    def _execute_answer_checker(self, question: str, answer: str) -> Tuple[bool, str]:
        """Internal method to execute Answer Checker agent operations."""
        # Create thread
        thread = self.project_client.agents.threads.create()
        
        # Create message
        message = self.project_client.agents.messages.create(
            thread_id=thread.id,
            role="user",
            content=f"Question: {question}\n\nCandidate Answer: {answer}\n\nPlease validate this answer for factual correctness, completeness, and consistency. Respond with 'VALID' if acceptable or 'INVALID: [reason]' if not."
        )
        
        # Create and process run
        run = self.project_client.agents.runs.create_and_process(
            thread_id=thread.id,
            agent_id=self.answer_checker_id
        )
        
        # Get messages
        messages = self.project_client.agents.messages.list(thread_id=thread.id)
        
        # Find the assistant's response
        for msg in messages:
            if msg.role == "assistant" and msg.content:
                response = msg.content[0].text.value
                if "VALID" in response.upper() and "INVALID" not in response.upper():
                    return True, response
                else:
                    return False, response
                    
        return False, "No response from Answer Checker"
            
    def validate_links(self, links: List[str]) -> Tuple[bool, List[str], str]:
        """Validate links using CURL and Link Checker agent."""
        try:
            # Use custom span for Link Checker agent
            if self.tracer:
                with self.tracer.start_as_current_span("link_checker_agent") as span:
                    span.set_attribute("agent.name", "Link Checker")
                    span.set_attribute("agent.operation", "validate_links")
                    span.set_attribute("links.count", len(links))
                    return self._execute_link_checker(links)
            else:
                return self._execute_link_checker(links)
                
        except Exception as e:
            self.logger.error(f"Error validating links: {e}")
            return False, [], f"Error: {e}"
    
    def _execute_link_checker(self, links: List[str]) -> Tuple[bool, List[str], str]:
        """Internal method to execute Link Checker agent operations."""
        import requests
        
        # First check: Must have at least one URL
        if not links or len(links) == 0:
            self.log_reasoning("Link Checker: No documentation URLs found - this is a failure condition")
            return False, [], "No documentation URLs provided. All answers must include source links."
        
        self.log_reasoning(f"Link Checker: Validating {len(links)} URLs")
        
        valid_links = []
        invalid_links = []
        
        # Check if links are reachable using requests
        for link in links:
            try:
                response = requests.head(link, timeout=10, allow_redirects=True)
                if response.status_code == 200:
                    valid_links.append(link)
                    self.log_reasoning(f"Link Checker: ✓ {link} (HTTP 200)")
                else:
                    invalid_links.append(f"{link} (HTTP {response.status_code})")
                    self.log_reasoning(f"Link Checker: ✗ {link} (HTTP {response.status_code})")
            except Exception as e:
                invalid_links.append(f"{link} (Error: {e})")
                self.log_reasoning(f"Link Checker: ✗ {link} (Error: {e})")
        
        # If we have at least one valid link, that's success
        if valid_links:
            if invalid_links:
                self.log_reasoning(f"Link Checker: Removed {len(invalid_links)} invalid URLs, keeping {len(valid_links)} valid ones")
                return True, valid_links, f"Found {len(valid_links)} valid links (removed {len(invalid_links)} invalid)"
            else:
                return True, valid_links, f"All {len(valid_links)} links are valid"
        else:
            return False, [], "No valid documentation URLs found after validation"
        
    def extract_links_and_clean(self, text: str) -> Tuple[str, List[str]]:
        """Extract URLs from text and return cleaned text and list of URLs."""
        import re
        
        self.log_reasoning(f"Extracting URLs from text (length: {len(text)})")
        self.log_reasoning(f"Text preview: {text[:200]}...")
        
        # Find URLs
        url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
        urls = re.findall(url_pattern, text)
        self.log_reasoning(f"URL regex found {len(urls)} URLs")
        
        # Remove URLs from text
        clean_text = re.sub(url_pattern, '', text)
        
        # Remove various citation formats
        clean_text = re.sub(r'\[\d+\]', '', clean_text)  # Remove [1], [2], etc.
        clean_text = re.sub(r'【[^】]*】', '', clean_text)  # Remove 【3:3†source】 style citations
        clean_text = re.sub(r'\(\d+\)', '', clean_text)  # Remove (1), (2), etc.
        clean_text = re.sub(r'\[\d+:\d+[^]]*\]', '', clean_text)  # Remove [3:3†source] style
        
        # Remove markdown formatting
        clean_text = re.sub(r'\*\*(.*?)\*\*', r'\1', clean_text)  # Remove **bold**
        clean_text = re.sub(r'\*(.*?)\*', r'\1', clean_text)  # Remove *italic*
        clean_text = re.sub(r'`(.*?)`', r'\1', clean_text)  # Remove `code`
        clean_text = re.sub(r'#{1,6}\s*', '', clean_text)  # Remove # headers
        
        # Remove numbered list formatting
        clean_text = re.sub(r'^\d+\.\s*\*\*[^*]*\*\*:\s*', '', clean_text, flags=re.MULTILINE)  # Remove "1. **Title**:"
        clean_text = re.sub(r'^\d+\.\s*', '', clean_text, flags=re.MULTILINE)  # Remove "1. "
        
        # Remove bullet points
        clean_text = re.sub(r'^\s*[-•]\s*', '', clean_text, flags=re.MULTILINE)
        
        # Remove "References:" and similar closing phrases at the end
        clean_text = re.sub(r'\s*References?:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more information,?\s*see:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more information,?\s*visit:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more details,?\s*see:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Learn more:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Learn more at:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*More information:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Additional resources:\s*$', '', clean_text, flags=re.IGNORECASE)
        
        # Clean up whitespace and line breaks
        clean_text = re.sub(r'\n\s*\n', ' ', clean_text)  # Replace multiple newlines with space
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()  # Clean up multiple spaces
        
        return clean_text, urls
        
    def on_import_excel_clicked(self):
        """Handle the Import From Excel button click."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            # Process Excel file in separate thread
            thread = threading.Thread(target=self.process_excel_file, args=(file_path,))
            thread.daemon = True
            thread.start()
            
    def process_excel_file(self, file_path: str):
        """Process an Excel file with questions."""
        try:
            self.log_reasoning(f"Processing Excel file: {file_path}")
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                self.create_agents()
            
            # Create temporary copy
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Copy original file
            import shutil
            shutil.copy2(file_path, temp_path)
            
            # Read Excel file
            excel_file = pd.ExcelFile(temp_path)
            context = self.context_entry.get().strip()
            char_limit = int(self.limit_entry.get()) if self.limit_entry.get().isdigit() else 2000
            max_retries = int(self.retries_entry.get()) if self.retries_entry.get().isdigit() else self.max_retries
            
            for sheet_name in excel_file.sheet_names:
                self.log_reasoning(f"Processing sheet: {sheet_name}")
                df = pd.read_excel(temp_path, sheet_name=sheet_name)
                
                # Use LLM to identify columns (reuse CLI logic)
                question_col, answer_col, docs_col = self.identify_columns_with_llm_cli(df)
                
                self.log_reasoning(f"LLM identified columns - Questions: {question_col}, Answers: {answer_col}, Docs: {docs_col}")
                
                # Skip sheet if no question or answer column found
                if not question_col or not answer_col:
                    self.log_reasoning(f"Skipping sheet '{sheet_name}' - missing required question or answer column")
                    continue
                
                # Process each question
                questions_processed = 0
                questions_attempted = 0
                for idx, row in df.iterrows():
                    if pd.notna(row[question_col]) and str(row[question_col]).strip():
                        question = str(row[question_col]).strip()
                        questions_attempted += 1
                        self.log_reasoning(f"Processing question {idx + 1}: {question[:50]}...")
                        
                        # Process question using the same workflow as single questions
                        success, answer, links = self.process_question_with_agents(question, context, char_limit, max_retries)
                        
                        if success:
                            # Update the answer column
                            df.at[idx, answer_col] = answer
                            
                            # Update documentation column only if it exists and we have links
                            if docs_col and links:
                                df.at[idx, docs_col] = '\n'.join(links)
                            # Leave documentation blank if no links or no docs column
                            
                            self.log_reasoning(f"Successfully processed question {idx + 1}")
                            questions_processed += 1
                        else:
                            self.log_reasoning(f"Failed to process question {idx + 1}: {answer}")
                            # Leave response blank on failure - don't write error messages  
                            # Leave documentation blank on failure - don't write error messages
                
                # Save updated sheet if we attempted any questions (regardless of success)
                if questions_attempted > 0:
                    # Use openpyxl directly to preserve formatting
                    from openpyxl import load_workbook
                    wb = load_workbook(temp_path)
                    ws = wb[sheet_name]
                    
                    # Update only the data, not the formatting
                    for idx, row in df.iterrows():
                        row_num = idx + 2  # +2 because Excel is 1-indexed and has header
                        if question_col and pd.notna(row[question_col]) and str(row[question_col]).strip():
                            # Find answer column index
                            for col_idx, col_name in enumerate(df.columns, 1):
                                if col_name == answer_col:
                                    cell = ws.cell(row=row_num, column=col_idx)
                                    if pd.notna(row[answer_col]) and str(row[answer_col]).strip():
                                        cell.value = str(row[answer_col])
                                elif col_name == docs_col and docs_col:
                                    cell = ws.cell(row=row_num, column=col_idx)
                                    if pd.notna(row[docs_col]) and str(row[docs_col]).strip():
                                        cell.value = str(row[docs_col])
                    
                    wb.save(temp_path)
                
                self.log_reasoning(f"Processed {questions_processed}/{questions_attempted} questions successfully in sheet '{sheet_name}'")
            
            # Ask user where to save
            self.root.after(0, lambda: self.save_processed_excel(temp_path))
                
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {e}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to process Excel file:\n{e}"))
            
    def save_processed_excel(self, temp_path: str):
        """Save the processed Excel file."""
        try:
            save_path = filedialog.asksaveasfilename(
                title="Save Processed Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if save_path:
                import shutil
                shutil.move(temp_path, save_path)
                messagebox.showinfo("Success", f"File processed and saved to:\n{save_path}")
            else:
                os.unlink(temp_path)
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")
            
    def identify_columns_with_llm(self, df: pd.DataFrame) -> Tuple[str, str, str]:
        """Use LLM to identify question, answer, and documentation columns."""
        try:
            # Create a prompt with column names and sample data
            column_info = []
            for col in df.columns:
                sample_data = df[col].dropna().head(3).tolist()
                column_info.append(f"Column '{col}': {sample_data}")
            
            prompt = f"""Analyze the following Excel columns and identify which column contains:
1. Questions (to be answered)
2. Expected answers (where AI responses should go)
3. Documentation/links (where reference links should go)

Columns:
{chr(10).join(column_info)}

Respond in this exact format:
Question Column: [column_name]
Answer Column: [column_name]
Documentation Column: [column_name]

If a column doesn't exist, suggest a name for it."""

            # Use the Question Answerer agent to analyze columns
            thread = self.project_client.agents.threads.create()
            message = self.project_client.agents.messages.create(
                thread_id=thread.id,
                role="user",
                content=prompt
            )
            
            run = self.project_client.agents.runs.create_and_process(
                thread_id=thread.id,
                agent_id=self.question_answerer_id
            )
            
            messages = self.project_client.agents.messages.list(thread_id=thread.id)
            
            # Parse the response
            for msg in messages:
                if msg.role == "assistant" and msg.content:
                    response = msg.content[0].text.value
                    question_col = self.extract_column_name(response, "Question Column:")
                    answer_col = self.extract_column_name(response, "Answer Column:")
                    docs_col = self.extract_column_name(response, "Documentation Column:")
                    
                    # Fallback to simple logic if parsing fails
                    if not question_col:
                        question_col = self.identify_question_column(df)
                    if not answer_col:
                        answer_col = self.identify_answer_column(df)
                    if not docs_col:
                        docs_col = self.identify_docs_column(df)
                    
                    return question_col, answer_col, docs_col
            
            # Fallback to simple logic
            return (self.identify_question_column(df), 
                   self.identify_answer_column(df), 
                   self.identify_docs_column(df))
            
        except Exception as e:
            self.logger.error(f"Error identifying columns with LLM: {e}")
            # Fallback to simple logic
            return (self.identify_question_column(df), 
                   self.identify_answer_column(df), 
                   self.identify_docs_column(df))
                   
    def extract_column_name(self, text: str, prefix: str) -> Optional[str]:
        """Extract column name from LLM response."""
        import re
        pattern = f"{re.escape(prefix)}\\s*(.+?)(?:\\n|$)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return None
        
    def process_question_with_agents(self, question: str, context: str, char_limit: int, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Process a single question using the three-agent workflow."""
        try:
            attempt = 1
            max_attempts = max_retries
            
            # Track valid links across retries for this question
            accumulated_valid_links = []
            
            while attempt <= max_attempts:
                # Step 1: Generate answer
                candidate_answer, doc_urls = self.generate_answer(question, context, char_limit)
                
                if not candidate_answer:
                    return False, "Question Answerer failed to generate an answer", []
                
                # Remove links and citations, save links for documentation
                clean_answer, text_links = self.extract_links_and_clean(candidate_answer)
                
                # Combine documentation URLs from run steps with any URLs found in text
                all_links = list(set(doc_urls + text_links))  # Remove duplicates
                
                # Check character limit
                if len(clean_answer) > char_limit:
                    attempt += 1
                    continue
                
                # Step 2: Validate answer
                answer_valid, answer_feedback = self.validate_answer(question, clean_answer)
                
                if not answer_valid:
                    attempt += 1
                    continue
                
                # Step 3: Validate links
                links_valid, valid_links, link_feedback = self.validate_links(all_links)
                
                # Accumulate any valid links found in this attempt
                if valid_links:
                    for link in valid_links:
                        if link not in accumulated_valid_links:
                            accumulated_valid_links.append(link)
                
                # Check if we have a valid answer and at least some valid links (current or accumulated)
                if not links_valid and not accumulated_valid_links:
                    # No valid links in current attempt and no accumulated links from previous attempts
                    attempt += 1
                    continue
                elif not links_valid and accumulated_valid_links:
                    # Current attempt has no valid links, but we have accumulated valid links from previous attempts
                    final_valid_links = accumulated_valid_links.copy()
                else:
                    # Current attempt has valid links
                    final_valid_links = accumulated_valid_links.copy()  # Use all accumulated links
                
                # All checks passed
                return True, clean_answer, final_valid_links
                
            # Max attempts reached
            return False, f"Failed to generate acceptable answer after {max_attempts} attempts", []
            
        except Exception as e:
            self.logger.error(f"Error processing question with agents: {e}")
            return False, f"Error: {e}", []
            
    def identify_columns_with_llm_cli(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """Use LLM to identify question, answer, and documentation columns for CLI mode."""
        try:
            # Create a prompt with column names and sample data
            column_info = []
            for col in df.columns:
                sample_data = df[col].dropna().head(3).tolist()
                # Convert to strings and truncate if too long
                sample_strings = [str(item)[:100] for item in sample_data]
                column_info.append(f"Column '{col}': {sample_strings}")
            
            prompt = f"""Analyze the following Excel columns and classify each column header as one of these types:
- QUESTION: Contains questions to be answered
- RESPONSE: Where AI responses/answers should be written
- DOCUMENTATION: Where reference links/documentation should be written  
- NONE: Not relevant for question answering

Columns:
{chr(10).join(column_info)}

Respond in this exact format:
Question Column: [column_name or NONE if no suitable column found]
Response Column: [column_name or NONE if no suitable column found]  
Documentation Column: [column_name or NONE if no suitable column found]

Only return existing column names. Do not suggest new column names."""

            # Use the Question Answerer agent to analyze columns
            thread = self.project_client.agents.threads.create()
            message = self.project_client.agents.messages.create(
                thread_id=thread.id,
                role="user",
                content=prompt
            )
            
            run = self.project_client.agents.runs.create_and_process(
                thread_id=thread.id,
                agent_id=self.question_answerer_id
            )
            
            messages = self.project_client.agents.messages.list(thread_id=thread.id)
            
            # Parse the response
            for msg in messages:
                if msg.role == "assistant" and msg.content:
                    response = msg.content[0].text.value
                    question_col = self.extract_column_name(response, "Question Column:")
                    answer_col = self.extract_column_name(response, "Response Column:")
                    docs_col = self.extract_column_name(response, "Documentation Column:")
                    
                    # Convert "NONE" to None
                    question_col = None if question_col and question_col.upper() == "NONE" else question_col
                    answer_col = None if answer_col and answer_col.upper() == "NONE" else answer_col
                    docs_col = None if docs_col and docs_col.upper() == "NONE" else docs_col
                    
                    # Validate that columns exist in the dataframe
                    if question_col and question_col not in df.columns:
                        question_col = None
                    if answer_col and answer_col not in df.columns:
                        answer_col = None  
                    if docs_col and docs_col not in df.columns:
                        docs_col = None
                    
                    return question_col, answer_col, docs_col
            
            # Fallback if LLM fails
            return None, None, None
            
        except Exception as e:
            self.logger.error(f"Error identifying columns with LLM: {e}")
            return None, None, None
    
    def extract_column_name(self, response: str, prefix: str) -> Optional[str]:
        """Extract column name from LLM response."""
        lines = response.split('\n')
        for line in lines:
            if line.strip().startswith(prefix):
                # Extract everything after the colon and clean it up
                part = line.split(':', 1)[1].strip()
                # Remove quotes and brackets if present
                part = part.strip('"\'[]')
                return part if part else None
        return None
        
    def run(self):
        """Start the application."""
        if not self.headless_mode:
            self.root.mainloop()
    
    def process_single_question_cli(self, question: str, context: str, char_limit: int, verbose: bool, max_retries: int = None) -> Tuple[bool, str, List[str]]:
        """Process a single question in CLI mode."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
        
        if self.tracer:
            with self.tracer.start_as_current_span("questionnaire_agent_cli") as span:
                span.set_attribute("interface.type", "CLI")
                span.set_attribute("question.length", len(question))
                span.set_attribute("question.preview", question[:100])
                span.set_attribute("context", context)
                span.set_attribute("char_limit", char_limit)
                span.set_attribute("verbose_mode", verbose)
                span.set_attribute("max_retries", max_retries)
                
                return self._execute_cli_workflow(question, context, char_limit, verbose, max_retries)
        else:
            return self._execute_cli_workflow(question, context, char_limit, verbose, max_retries)
    
    def _execute_cli_workflow(self, question: str, context: str, char_limit: int, verbose: bool, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Execute the CLI workflow using the same multi-agent approach as GUI."""
        try:
            if verbose:
                print("Starting question processing...")
            
            # Store the current CLI output buffer
            original_cli_output = self.cli_output.copy()
            self.cli_output.clear()
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                self.create_agents()
            
            # Use the same workflow as GUI mode
            success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
            
            # Extract results from CLI output (the workflow populates this in headless mode)
            if verbose:
                for line in self.cli_output:
                    print(line)
            
            return success, answer, links
                
        except Exception as e:
            if verbose:
                print(f"Error: {e}")
            return False, f"Error: {e}", []
        finally:
            # Restore original CLI output
            self.cli_output = original_cli_output
    
    
    def process_excel_file_cli(self, input_path: str, output_path: str, context: str, char_limit: int, verbose: bool, max_retries: int = None) -> bool:
        """Process an Excel file in CLI mode."""
        scenario = "questionnaire_agent_excel_processing"
        
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
        
        if self.tracer:
            with self.tracer.start_as_current_span(scenario) as span:
                span.set_attribute("input_file", input_path)
                span.set_attribute("output_file", output_path)
                span.set_attribute("context", context)
                span.set_attribute("char_limit", char_limit)
                span.set_attribute("verbose_mode", verbose)
                span.set_attribute("max_retries", max_retries)
                
                return self._process_excel_file_internal(input_path, output_path, context, char_limit, verbose, span, max_retries)
        else:
            return self._process_excel_file_internal(input_path, output_path, context, char_limit, verbose, None, max_retries)
    
    def _process_excel_file_internal(self, input_path: str, output_path: str, context: str, char_limit: int, verbose: bool, span=None, max_retries: int = None) -> bool:
        """Internal method for Excel file processing with tracing."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
            
        try:
            if verbose:
                print(f"Processing Excel file: {input_path}")
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                if span and self.tracer:
                    with self.tracer.start_as_current_span("create_agents") as agent_span:
                        agent_span.set_attribute("operation", "create_all_agents")
                        self.create_agents()
                else:
                    self.create_agents()
            
            # Read Excel file
            excel_file = pd.ExcelFile(input_path)
            
            # Create output file by copying input
            import shutil
            shutil.copy2(input_path, output_path)
            
            total_sheets = len(excel_file.sheet_names)
            if span:
                span.set_attribute("sheets.total_count", total_sheets)
                span.set_attribute("sheets.names", excel_file.sheet_names)
            
            for sheet_index, sheet_name in enumerate(excel_file.sheet_names, 1):
                if verbose:
                    print(f"Processing sheet: {sheet_name}")
                
                if span and self.tracer:
                    with self.tracer.start_as_current_span("process_excel_sheet") as sheet_span:
                        sheet_span.set_attribute("sheet.name", sheet_name)
                        sheet_span.set_attribute("sheet.index", sheet_index)
                        sheet_span.set_attribute("sheet.total", total_sheets)
                        
                        df = pd.read_excel(output_path, sheet_name=sheet_name)
                        sheet_span.set_attribute("sheet.row_count", len(df))
                        sheet_span.set_attribute("sheet.column_count", len(df.columns))
                        
                        self._process_excel_sheet(df, sheet_name, output_path, context, char_limit, verbose, sheet_span, max_retries)
                else:
                    df = pd.read_excel(output_path, sheet_name=sheet_name)
                    self._process_excel_sheet(df, sheet_name, output_path, context, char_limit, verbose, None, max_retries)
            
            if verbose:
                print(f"Excel processing completed. Results saved to: {output_path}")
            
            if span:
                span.set_attribute("success", True)
            
            return True
                
        except Exception as e:
            error_msg = f"Error processing Excel file: {e}"
            self.logger.error(error_msg)
            if verbose:
                print(error_msg)
            
            if span:
                span.set_attribute("success", False)
                span.set_attribute("error.message", str(e))
                span.set_attribute("error.type", type(e).__name__)
                span.set_status(trace.Status(trace.StatusCode.ERROR, str(e)))
            
            return False
    
    def _process_excel_sheet(self, df: pd.DataFrame, sheet_name: str, output_path: str, context: str, char_limit: int, verbose: bool, span=None, max_retries: int = None):
        """Process a single Excel sheet with tracing."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
            
        # Use LLM to identify columns
        question_col, answer_col, docs_col = self.identify_columns_with_llm_cli(df)
        
        if verbose:
            print(f"LLM identified columns - Questions: {question_col}, Answers: {answer_col}, Docs: {docs_col}")
        
        if span:
            span.set_attribute("columns.question", question_col or "none")
            span.set_attribute("columns.answer", answer_col or "none") 
            span.set_attribute("columns.docs", docs_col or "none")
        
        # Skip sheet if no question or answer column found
        if not question_col or not answer_col:
            if verbose:
                print(f"Skipping sheet '{sheet_name}' - missing required question or answer column")
            if span:
                span.set_attribute("sheet.skipped", True)
                span.set_attribute("skip_reason", "missing_required_columns")
            return
        
        # Process each question
        questions_processed = 0
        questions_attempted = 0
        for idx, row in df.iterrows():
            if pd.notna(row[question_col]) and str(row[question_col]).strip():
                question = str(row[question_col]).strip()
                questions_attempted += 1
                if verbose:
                    print(f"Processing question {idx + 1}: {question[:50]}...")
                
                # Process question using CLI workflow
                success, answer, links = self.process_single_question_cli(question, context, char_limit, False, max_retries)
                
                if success:
                    # Update the answer column
                    df.at[idx, answer_col] = answer
                    
                    # Update documentation column only if it exists and we have links
                    if docs_col and links:
                        df.at[idx, docs_col] = '\n'.join(links)
                    # Leave documentation blank if no links or no docs column
                    
                    if verbose:
                        print(f"Successfully processed question {idx + 1}")
                    questions_processed += 1
                else:
                    if verbose:
                        print(f"Failed to process question {idx + 1}: {answer}")
                    # Leave response blank on failure - don't write error messages
                    # Leave documentation blank on failure - don't write error messages
        
        if span:
            span.set_attribute("questions.attempted", questions_attempted)
            span.set_attribute("questions.processed", questions_processed)
            span.set_attribute("questions.success_rate", questions_processed / questions_attempted if questions_attempted > 0 else 0)
        
        # Save updated sheet if we attempted any questions (regardless of success)
        if questions_attempted > 0:
            # Use openpyxl directly to preserve formatting
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb[sheet_name]
            
            # Update only the data, not the formatting
            for idx, row in df.iterrows():
                row_num = idx + 2  # +2 because Excel is 1-indexed and has header
                if question_col and pd.notna(row[question_col]) and str(row[question_col]).strip():
                    # Find answer column index
                    for col_idx, col_name in enumerate(df.columns, 1):
                        if col_name == answer_col:
                            cell = ws.cell(row=row_num, column=col_idx)
                            if pd.notna(row[answer_col]) and str(row[answer_col]).strip():
                                cell.value = str(row[answer_col])
                        elif col_name == docs_col and docs_col:
                            cell = ws.cell(row=row_num, column=col_idx)
                            if pd.notna(row[docs_col]) and str(row[docs_col]).strip():
                                cell.value = str(row[docs_col])
            
            wb.save(output_path)
        
        if verbose:
            print(f"Processed {questions_processed}/{questions_attempted} questions successfully in sheet '{sheet_name}'")


def create_cli_parser():
    """Create command line argument parser."""
    parser = argparse.ArgumentParser(
        description="Questionnaire Multiagent - AI question answering with fact-checking and link validation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # Single question
  python question_answerer.py --question "Does your service offer video generative AI?" --context "Microsoft Azure AI" --char-limit 2000
  
  # Excel processing
  python question_answerer.py --import-excel questions.xlsx --output-excel processed.xlsx --context "Microsoft Azure AI" --verbose"""
    )
    
    parser.add_argument('-q', '--question', type=str, help='The natural-language question to ask')
    parser.add_argument('-c', '--context', type=str, default='Microsoft Azure AI', help='Context or topic string to bias the question answering (default: "Microsoft Azure AI")')
    parser.add_argument('--char-limit', type=int, default=2000, help='Integer character limit for the final answer (default: 2000)')
    parser.add_argument('--max-retries', type=int, default=10, help='Maximum number of retries for answer generation (default: 10)')
    parser.add_argument('--import-excel', type=str, metavar='PATH', help='Path to an Excel file to process in batch')
    parser.add_argument('--output-excel', type=str, metavar='PATH', help='Path where the processed Excel file will be written')
    parser.add_argument('--verbose', action='store_true', default=True, help='Enable verbose/reasoning log output (default: True)')
    
    return parser


def main():
    """Main entry point."""
    parser = create_cli_parser()
    
    # Check if any arguments were provided
    if len(sys.argv) == 1:
        # No arguments - run in GUI mode
        try:
            app = QuestionnaireAgentUI(headless_mode=False)
            app.run()
        except Exception as e:
            print(f"Failed to start application: {e}")
            try:
                messagebox.showerror("Startup Error", f"Failed to start application:\n{e}")
            except:
                pass
    else:
        # Arguments provided - run in CLI mode
        args = parser.parse_args()
        
        # Validate arguments
        if not args.question and not args.import_excel:
            print("Error: Either --question or --import-excel must be provided")
            parser.print_help()
            sys.exit(1)
        
        if args.import_excel and not args.output_excel:
            # Generate default output filename
            input_path = Path(args.import_excel)
            args.output_excel = str(input_path.parent / f"{input_path.stem}.answered.xlsx")
        
        try:
            app = QuestionnaireAgentUI(headless_mode=True, max_retries=args.max_retries)
            
            if args.question:
                # Process single question
                success, answer, links = app.process_single_question_cli(
                    args.question, args.context, args.char_limit, args.verbose, args.max_retries
                )
                
                if success:
                    print("\n=== ANSWER ===")
                    print(answer)
                    if links:
                        print("\n=== DOCUMENTATION LINKS ===")
                        for link in links:
                            print(f"• {link}")
                    else:
                        print("\n=== DOCUMENTATION LINKS ===")
                        print("No documentation links found")
                    sys.exit(0)
                else:
                    print(f"Error: {answer}")
                    sys.exit(1)
            
            elif args.import_excel:
                # Process Excel file
                if not os.path.exists(args.import_excel):
                    print(f"Error: Excel file not found: {args.import_excel}")
                    sys.exit(1)
                
                success = app.process_excel_file_cli(
                    args.import_excel, args.output_excel, args.context, args.char_limit, args.verbose, args.max_retries
                )
                
                if success:
                    print(f"\nExcel processing completed successfully. Results saved to: {args.output_excel}")
                    sys.exit(0)
                else:
                    print("Error: Excel processing failed")
                    sys.exit(1)
        
        except KeyboardInterrupt:
            print("\nOperation cancelled by user")
            sys.exit(1)
        except Exception as e:
            print(f"Error: {e}")
            sys.exit(1)


if __name__ == "__main__":
    main()